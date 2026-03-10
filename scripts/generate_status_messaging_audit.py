#!/usr/bin/env python3
from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from collections import defaultdict
from typing import List, Dict, Tuple

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font

ROOT = Path(__file__).resolve().parents[1]
SRC = ROOT / "Sources" / "Ledger"
OUT = ROOT / "docs" / "status-messaging-audit-v1.1.xlsx"


@dataclass
class Callsite:
    template: str
    file: str
    line: int
    symbol: str
    surface: str
    trigger: str
    snippet: str
    planned_action: str = "No Change"


def read_files() -> List[Path]:
    return sorted(SRC.rglob("*.swift"))


def nearest_symbol(lines: List[str], idx: int) -> str:
    for i in range(idx, -1, -1):
        line = lines[i].strip()
        m = re.search(r"\bfunc\s+([A-Za-z0-9_]+)", line)
        if m:
            return f"func {m.group(1)}"
        m = re.search(r"\b(var|let)\s+([A-Za-z0-9_]+)", line)
        if m and (":" in line or "=" in line):
            return f"{m.group(1)} {m.group(2)}"
        m = re.search(r"\bstruct\s+([A-Za-z0-9_]+)", line)
        if m:
            return f"struct {m.group(1)}"
        m = re.search(r"\bclass\s+([A-Za-z0-9_]+)", line)
        if m:
            return f"class {m.group(1)}"
    return "(file scope)"


def extract_string_literals(text: str) -> List[str]:
    result = []
    for m in re.finditer(r'"([^"\\]*(?:\\.[^"\\]*)*)"', text):
        value = m.group(1).strip()
        if value:
            result.append(value)
    return result


def classify_flow(file: str, symbol: str, template: str) -> str:
    low = f"{file} {symbol} {template}".lower()
    if any(k in low for k in ["apply", "restore", "backup", "undo", "redo"]):
        return "Apply/Restore"
    if any(k in low for k in ["import", "csv", "gpx", "reference", "eos", "previewerror"]):
        return "Import"
    if "preset" in low:
        return "Presets"
    if any(k in low for k in ["open", "finder", "reveal", "folder", "pinned", "recent", "sidebar"]):
        return "Navigation/Open/Reveal"
    if any(k in low for k in ["quit", "launch", "exiftool", "settings", "permission", "startup"]):
        return "App Lifecycle/Settings"
    return "General"


def classify_trigger(file: str, symbol: str, template: str) -> str:
    low = f"{file} {symbol} {template}".lower()
    if "runmodal" in low or "alert" in low:
        return "User decision prompt"
    if "import" in low:
        return "Import flow"
    if any(k in low for k in ["apply", "restore"]):
        return "Apply/restore flow"
    if "preset" in low:
        return "Preset flow"
    if any(k in low for k in ["open", "finder", "reveal", "sidebar"]):
        return "Navigation action"
    if any(k in low for k in ["quit", "launch", "exiftool"]):
        return "App lifecycle"
    return "UI action"


def blocking_level(surface: str, template: str) -> str:
    low = template.lower()
    if surface == "Modal":
        if any(k in low for k in ["quit", "discard", "apply", "restore", "failed", "requires"]):
            return "Blocks app flow"
        return "Blocks current flow"
    if surface in {"Inline Sheet", "View-State"}:
        return "Blocks current flow"
    return "Non-blocking"


def category(surface: str, template: str) -> str:
    low = template.lower()
    if surface == "Modal" and any(k in low for k in ["?", "choose", "apply", "discard", "quit"]):
        return "Decision Prompt"
    if any(k in low for k in ["couldn", "failed", "error", "invalid", "no "]):
        return "Error"
    if any(k in low for k in ["warning", "select", "disabled", "not found"]):
        return "Warning"
    if any(k in low for k in ["saved", "updated", "deleted", "applied", "restored", "refreshed", "loaded", "added"]):
        return "Success"
    return "Info"


def modal_justification(surface: str, template: str) -> str:
    if surface != "Modal":
        return "None"
    low = template.lower()
    if any(k in low for k in ["discard", "quit", "undo", "can't be undone"]):
        return "Data-loss risk"
    if any(k in low for k in ["apply", "restore", "write"]):
        return "Irreversible"
    if any(k in low for k in ["requires", "permission", "could not be found"]):
        return "Permission/blocking failure"
    return "None"

def default_rewrite(template: str) -> str:
    rewritten = template
    rewritten = rewritten.replace("Staged", "Prepared")
    rewritten = rewritten.replace("staged", "prepared")
    rewritten = rewritten.replace("Apply failed", "Couldn’t apply metadata changes.")
    rewritten = rewritten.replace("Restore failed", "Couldn’t restore metadata.")
    rewritten = rewritten.replace("Export Failed", "Couldn’t export ExifTool CSV.")
    return rewritten


def default_decision(surface: str, template: str) -> tuple[str, str, str, str, str]:
    low = template.lower()
    proposed_surface = surface
    change_type = "Text only"
    decision = "Keep as Status Bar" if surface == "Status Bar" else "Keep but Rewrite"
    if surface == "Modal":
        decision = "Keep but Rewrite"
    if surface == "Inline Sheet":
        if "conflict" in low and "need resolution" in low:
            decision = "Promote to Modal"
            proposed_surface = "Modal"
            change_type = "Surface change"
        else:
            decision = "Keep but Rewrite"
    if "staged" in low:
        decision = "Keep but Rewrite"
    return ("Yes", decision, proposed_surface, change_type, default_rewrite(template))


def extract_callsites() -> List[Callsite]:
    out: List[Callsite] = []
    for path in read_files():
        rel = str(path.relative_to(ROOT))
        lines = path.read_text(encoding="utf-8").splitlines()
        for i, line in enumerate(lines):
            stripped = line.strip()
            symbol = nearest_symbol(lines, i)

            # Status bar messaging
            if "statusMessage" in stripped and "=" in stripped:
                for s in extract_string_literals(stripped):
                    out.append(Callsite(s, rel, i + 1, symbol, "Status Bar", classify_trigger(rel, symbol, s), stripped))

            if "setStatusMessage(" in stripped:
                args = stripped.split("setStatusMessage(", 1)[1]
                for s in extract_string_literals(args):
                    out.append(Callsite(s, rel, i + 1, symbol, "Status Bar", classify_trigger(rel, symbol, s), stripped))

            # Import inline errors
            if "previewError" in stripped and "=" in stripped:
                for s in extract_string_literals(stripped):
                    out.append(Callsite(s, rel, i + 1, symbol, "Inline Sheet", classify_trigger(rel, symbol, s), stripped))

            # SwiftUI alerts
            if ".alert(" in stripped:
                for s in extract_string_literals(stripped):
                    out.append(Callsite(s, rel, i + 1, symbol, "Modal", classify_trigger(rel, symbol, s), stripped))

            # NSAlert texts
            if "alert.messageText" in stripped or "alert.informativeText" in stripped:
                for s in extract_string_literals(stripped):
                    out.append(Callsite(s, rel, i + 1, symbol, "Modal", classify_trigger(rel, symbol, s), stripped))

            # NSOpenPanel prompt label
            if ".prompt" in stripped and "=" in stripped:
                for s in extract_string_literals(stripped):
                    out.append(Callsite(s, rel, i + 1, symbol, "Modal", classify_trigger(rel, symbol, s), stripped))

            # Shared blocking modal helper
            if "presentBlockingWarning(" in stripped or "presentBlockingImportAlert(" in stripped:
                for s in extract_string_literals(stripped):
                    out.append(Callsite(s, rel, i + 1, symbol, "Modal", classify_trigger(rel, symbol, s), stripped))

            # View-state enum error surfaces
            if "enumerationError(" in stripped and "return" in stripped:
                out.append(Callsite("Folder enumeration error state", rel, i + 1, symbol, "View-State", classify_trigger(rel, symbol, "enumeration error"), stripped))

    # remove overly noisy placeholders
    filtered = []
    seen = set()
    for c in out:
        key = (c.template, c.file, c.line, c.surface)
        if c.template in {"Ready", "", "OK", "Cancel", "Open"}:
            continue
        if key in seen:
            continue
        seen.add(key)
        filtered.append(c)
    return filtered


def make_batches(flows: List[str]) -> Dict[str, str]:
    counters: Dict[str, int] = defaultdict(int)
    result: Dict[str, str] = {}
    for flow in flows:
        prefix = flow.upper().replace("/", "-").replace(" ", "-")
        counters[prefix] += 1
        result[flow] = f"FLOW-{prefix}-{counters[prefix]:02d}"
    return result


def build_workbook(callsites: List[Callsite]) -> None:
    wb = Workbook()
    ws_audit = wb.active
    ws_audit.title = "Audit"
    ws_calls = wb.create_sheet("Callsites")

    audit_headers = [
        "Message ID", "User Flow", "Current Surface", "Message Template", "Blocking Level",
        "Necessary?", "Decision", "Proposed Surface", "Modal Justification", "Rewrite Proposal",
        "Implementation Batch", "Primary Owner File", "Change Type", "Risk", "Test Scenario ID",
        "Implementation Status", "Notes", "Editable Final Text"
    ]
    calls_headers = [
        "Message ID", "File", "Line", "Symbol/Method", "Current Surface", "Trigger Context",
        "Code Snippet (short)", "Planned Action"
    ]

    ws_audit.append(audit_headers)
    ws_calls.append(calls_headers)

    grouped: Dict[Tuple[str, str], List[Callsite]] = defaultdict(list)
    for c in callsites:
        grouped[(c.surface, c.template)].append(c)

    ids = {}
    for idx, key in enumerate(sorted(grouped.keys(), key=lambda x: (x[0], x[1].lower())), start=1):
        ids[key] = f"MSG-{idx:04d}"

    flow_batch_counter: Dict[str, int] = defaultdict(int)
    flow_test_counter: Dict[str, int] = defaultdict(int)

    for key in sorted(grouped.keys(), key=lambda x: ids[x]):
        surface, template = key
        rows = grouped[key]
        first = rows[0]
        flow = classify_flow(first.file, first.symbol, template)

        flow_key = flow.upper().replace("/", "-").replace(" ", "-")
        flow_batch_counter[flow_key] += 1
        impl_batch = f"FLOW-{flow_key}-{flow_batch_counter[flow_key]:02d}"
        flow_test_counter[flow_key] += 1
        test_id = f"TS-{flow_key}-{flow_test_counter[flow_key]:03d}"

        ws_audit.append([
            ids[key],
            flow,
            surface,
            template,
            blocking_level(surface, template),
            default_decision(surface, template)[0],
            default_decision(surface, template)[1],
            default_decision(surface, template)[2],
            modal_justification(surface, template),
            default_decision(surface, template)[4],
            impl_batch,
            first.file,
            default_decision(surface, template)[3],
            "Med",
            test_id,
            "In Progress",
            "",
            default_decision(surface, template)[4],
        ])

        for r in rows:
            ws_calls.append([
                ids[key],
                r.file,
                r.line,
                r.symbol,
                r.surface,
                r.trigger,
                r.snippet[:180],
                r.planned_action,
            ])

    for ws in (ws_audit, ws_calls):
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    # Dropdown validations
    dv_necessary = DataValidation(type="list", formula1='"Yes,No,Unsure"')
    dv_decision = DataValidation(type="list", formula1='"Keep as Status Bar,Promote to Modal,Demote to Status,Remove,Keep but Rewrite"')
    dv_surface = DataValidation(type="list", formula1='"Status Bar,Modal,Inline Sheet,View-State,N/A"')
    dv_modal = DataValidation(type="list", formula1='"Destructive,Irreversible,Data-loss risk,Permission/blocking failure,None"')
    dv_change = DataValidation(type="list", formula1='"Text only,Surface change,Remove,Refactor"')
    dv_risk = DataValidation(type="list", formula1='"Low,Med,High"')
    dv_impl = DataValidation(type="list", formula1='"Open,In Progress,Done,Blocked"')

    ws_audit.add_data_validation(dv_necessary)
    ws_audit.add_data_validation(dv_decision)
    ws_audit.add_data_validation(dv_surface)
    ws_audit.add_data_validation(dv_modal)
    ws_audit.add_data_validation(dv_change)
    ws_audit.add_data_validation(dv_risk)
    ws_audit.add_data_validation(dv_impl)

    max_row = ws_audit.max_row
    dv_necessary.add(f"F2:F{max_row}")
    dv_decision.add(f"G2:G{max_row}")
    dv_surface.add(f"H2:H{max_row}")
    dv_modal.add(f"I2:I{max_row}")
    dv_change.add(f"M2:M{max_row}")
    dv_risk.add(f"N2:N{max_row}")
    dv_impl.add(f"P2:P{max_row}")

    # widths
    widths_a = {
        "A": 12, "B": 24, "C": 15, "D": 58, "E": 18, "F": 10, "G": 26, "H": 16,
        "I": 28, "J": 32, "K": 22, "L": 46, "M": 16, "N": 10, "O": 16, "P": 16,
        "Q": 22, "R": 58,
    }
    widths_c = {"A": 12, "B": 52, "C": 8, "D": 28, "E": 15, "F": 20, "G": 64, "H": 16}
    for k, v in widths_a.items():
        ws_audit.column_dimensions[k].width = v
    for k, v in widths_c.items():
        ws_calls.column_dimensions[k].width = v

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)


if __name__ == "__main__":
    callsites = extract_callsites()
    build_workbook(callsites)
    print(f"Generated {OUT}")
    print(f"Callsites: {len(callsites)}")
